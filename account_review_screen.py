# ui/account_review_screen.py
"""
Account Review Screen - Review accounts with transaction details before mapping.
"""

import customtkinter as ctk
from typing import List, Dict
import tkinter as tk
from tkinter import ttk


class SimpleAccountReviewScreen(ctk.CTkFrame):
    """
    Review Screen after Name Confirmation.
    Shows all accounts with summary and allows viewing transaction details.
    """
    
    def __init__(self, master, confirmed_names: List[str], name_aliases: Dict[str, str],
                 bank_statement_df, on_complete):
        super().__init__(master)
        
        self.confirmed_names = list(confirmed_names) if confirmed_names else []
        self.name_aliases = dict(name_aliases) if name_aliases else {}
        self.bank_statement_df = bank_statement_df
        self.on_complete_callback = on_complete
        
        self.status_label = None
        self.tree = None
        self.search_var = None
        
        # Group transactions
        self.account_transactions = self._group_transactions()
        
        # Track edited names
        self.edited_names = {}
        for name in self.account_transactions.keys():
            self.edited_names[name] = name
        
        self._create_ui()
        
        print(f"Account Review Screen initialized")
        print(f"Confirmed names: {self.confirmed_names}")
        print(f"Aliases: {self.name_aliases}")
        print(f"Account transactions: {list(self.account_transactions.keys())}")

    def _group_transactions(self) -> Dict[str, List[Dict]]:
        """Group transactions by account"""
        account_trans = {}
        
        if self.bank_statement_df is None or self.bank_statement_df.empty:
            print("WARNING: bank_statement_df is None or empty")
            return account_trans
        
        for idx, row in self.bank_statement_df.iterrows():
            extracted_name = row.get('EXTRACTED_NAME', '')
            final_name = row.get('FINAL_NAME', '')
            
            # Use final_name if available, otherwise extracted_name
            name_to_use = final_name if final_name else extracted_name
            
            # Resolve aliases
            primary_name = None
            
            if name_to_use in self.name_aliases:
                primary_name = self.name_aliases[name_to_use]
            elif extracted_name in self.name_aliases:
                primary_name = self.name_aliases[extracted_name]
            else:
                primary_name = name_to_use
            
            # Skip truly invalid names
            if not primary_name or str(primary_name).strip() == '' or str(primary_name).upper() in ['UNKNOWN', 'NONE', 'NAN']:
                continue
            
            if primary_name not in account_trans:
                account_trans[primary_name] = []
            
            debit = float(row.get('DEBIT', 0)) if row.get('DEBIT', 0) else 0
            credit = float(row.get('CREDIT', 0)) if row.get('CREDIT', 0) else 0
            
            txn_date = row.get('DATE', '')
            description = row.get('DESCRIPTION', '')
            ref_no = row.get('REFERENCE', '')
            balance = row.get('BALANCE', '')
            
            account_trans[primary_name].append({
                'date': str(txn_date),
                'description': str(description),
                'ref_no': str(ref_no),
                'debit': debit,
                'credit': credit,
                'balance': str(balance),
                'row_index': idx
            })
        
        print(f"Grouped transactions for {len(account_trans)} accounts")
        for name, txns in account_trans.items():
            print(f"  {name}: {len(txns)} transactions")
        
        return account_trans

    def _create_ui(self):
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        # Header
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.grid(row=0, column=0, pady=15, padx=20, sticky="ew")
        
        ctk.CTkLabel(
            header_frame,
            text="Review Extracted Accounts",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(side="left")
        
        ctk.CTkLabel(
            header_frame,
            text=f"{len(self.account_transactions)} accounts found",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        ).pack(side="right")
        
        # Toolbar
        toolbar_frame = ctk.CTkFrame(self, fg_color=("gray85", "gray25"))
        toolbar_frame.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="ew")
        
        ctk.CTkLabel(toolbar_frame, text="Search:").pack(side="left", padx=10)
        
        self.search_var = ctk.StringVar()
        search_entry = ctk.CTkEntry(toolbar_frame, textvariable=self.search_var, width=200)
        search_entry.pack(side="left", padx=5)
        search_entry.bind('<KeyRelease>', self._filter_table)
        
        ctk.CTkButton(
            toolbar_frame,
            text="Clear",
            width=60,
            fg_color="gray",
            command=self._clear_search
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            toolbar_frame,
            text="Export to Excel",
            width=120,
            fg_color="#2196F3",
            command=self._export_to_excel
        ).pack(side="right", padx=10)
        
        ctk.CTkButton(
            toolbar_frame,
            text="Refresh",
            width=80,
            fg_color="gray",
            command=self._refresh_table
        ).pack(side="right", padx=5)
        
        # Table
        self._create_table()
        
        # Buttons - FIXED: Larger, more visible buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.grid(row=3, column=0, padx=20, pady=15, sticky="ew")
        
        ctk.CTkButton(
            btn_frame,
            text="← Back to Name Confirmation",
            fg_color="gray",
            width=220,
            height=40,
            command=self._go_back
        ).pack(side="left")
        
        ctk.CTkButton(
            btn_frame,
            text="Continue to Mapping →",
            fg_color="green",
            hover_color="#2E7D32",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=220,
            height=45,
            command=self._continue
        ).pack(side="right")

    def _create_table(self):
        """Create treeview table"""
        table_frame = ctk.CTkFrame(self)
        table_frame.grid(row=2, column=0, padx=20, pady=5, sticky="nsew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        tree_container = tk.Frame(table_frame, bg='#2b2b2b')
        tree_container.grid(row=0, column=0, sticky="nsew")
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        columns = (
            'account_name', 'txn_count', 'receipt_count', 'payment_count',
            'total_receipts', 'total_payments', 'net_amount', 'sample_desc'
        )
        
        self.tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show='headings',
            selectmode='browse'
        )
        
        col_config = [
            ('account_name', 'Account Name', 200, tk.W),
            ('txn_count', 'Total Txns', 80, tk.CENTER),
            ('receipt_count', 'Receipts', 70, tk.CENTER),
            ('payment_count', 'Payments', 70, tk.CENTER),
            ('total_receipts', 'Total Receipts', 110, tk.E),
            ('total_payments', 'Total Payments', 110, tk.E),
            ('net_amount', 'Net Amount', 110, tk.E),
            ('sample_desc', 'Sample Description', 250, tk.W)
        ]
        
        for col_id, heading, width, anchor in col_config:
            self.tree.heading(col_id, text=heading, anchor=anchor)
            self.tree.column(col_id, width=width, anchor=anchor, minwidth=50)
        
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Style
        style = ttk.Style()
        available_themes = style.theme_names()
        if 'clam' in available_themes:
            style.theme_use('clam')
        
        style.configure("Treeview",
                       background="#2d2d2d",
                       foreground="#e0e0e0",
                       fieldbackground="#2d2d2d",
                       font=('Segoe UI', 10),
                       rowheight=28)
        style.configure("Treeview.Heading",
                       background="#4a4a4a",
                       foreground="white",
                       font=('Segoe UI', 10, 'bold'),
                       relief="flat",
                       borderwidth=1)
        style.map('Treeview',
                 background=[('selected', '#1976D2')],
                 foreground=[('selected', 'white')])
        
        self.tree.bind('<Double-1>', self._on_double_click)
        
        self.status_label = ctk.CTkLabel(
            table_frame, 
            text="", 
            font=ctk.CTkFont(size=10),
            text_color="gray"
        )
        self.status_label.grid(row=2, column=0, sticky="w", pady=5)
        
        self._populate_table()

    def _populate_table(self, filter_text=""):
        """Populate table with account data"""
        if self.tree is None:
            return
            
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        row_num = 0
        for account_name in sorted(self.account_transactions.keys()):
            transactions = self.account_transactions.get(account_name, [])
            
            if filter_text and filter_text.lower() not in account_name.lower():
                continue
            
            if not transactions:
                continue
            
            total_receipt = sum(t['credit'] for t in transactions if t['credit'] > 0)
            total_payment = sum(t['debit'] for t in transactions if t['debit'] > 0)
            receipt_count = sum(1 for t in transactions if t['credit'] > 0)
            payment_count = sum(1 for t in transactions if t['debit'] > 0)
            net_amount = total_receipt - total_payment
            
            sample_desc = transactions[0]['description'][:50] if transactions[0]['description'] else ""
            
            values = (
                account_name,
                len(transactions),
                receipt_count if receipt_count > 0 else "-",
                payment_count if payment_count > 0 else "-",
                self._format_currency(total_receipt) if total_receipt > 0 else "-",
                self._format_currency(total_payment) if total_payment > 0 else "-",
                self._format_currency(net_amount),
                sample_desc
            )
            
            if row_num % 2 == 0:
                tag = 'even'
            else:
                tag = 'odd'
            
            self.tree.insert('', tk.END, values=values, tags=(tag,))
            row_num += 1
        
        self.tree.tag_configure('even', background='#2d2d2d')
        self.tree.tag_configure('odd', background='#353535')
        
        self._update_status()

    def _filter_table(self, event=None):
        """Filter table"""
        if self.search_var:
            filter_text = self.search_var.get()
            self._populate_table(filter_text)

    def _clear_search(self):
        """Clear search"""
        if self.search_var:
            self.search_var.set("")
            self._populate_table()

    def _update_status(self):
        """Update status bar"""
        if self.status_label is None or self.tree is None:
            return
        
        total = len(self.tree.get_children())
        total_accounts = len(self.account_transactions)
        self.status_label.configure(
            text=f"Showing {total} of {total_accounts} accounts"
        )

    def _format_currency(self, amount):
        """Format currency"""
        if amount == 0:
            return "-"
        
        abs_amount = abs(amount)
        if abs_amount >= 10000000:
            formatted = f"{abs_amount / 10000000:,.2f} Cr"
        elif abs_amount >= 100000:
            formatted = f"{abs_amount / 100000:,.2f} L"
        else:
            formatted = f"{abs_amount:,.0f}"
        
        if amount < 0:
            return f"-{formatted}"
        return formatted

    def _on_double_click(self, event):
        """Handle double-click"""
        if self.tree is None:
            return
            
        selected = self.tree.selection()
        if not selected:
            return
        
        item = self.tree.item(selected[0])
        account_name = item['values'][0]
        self._view_details(account_name)

    def _view_details(self, account_name):
        """Show transaction details popup"""
        transactions = self.account_transactions.get(account_name, [])
        
        if not transactions:
            return
        
        popup = ctk.CTkToplevel(self)
        popup.title(f"Transactions: {account_name}")
        popup.geometry("950x500")
        popup.grab_set()
        
        # Header
        header_frame = ctk.CTkFrame(popup, fg_color=("gray80", "gray20"))
        header_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(
            header_frame,
            text=account_name,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#4CAF50"
        ).pack(side="left", padx=20, pady=10)
        
        total_receipt = sum(t['credit'] for t in transactions if t['credit'] > 0)
        total_payment = sum(t['debit'] for t in transactions if t['debit'] > 0)
        
        stats_text = f"{len(transactions)} transactions"
        if total_receipt > 0:
            stats_text += f" | Receipts: {self._format_currency(total_receipt)}"
        if total_payment > 0:
            stats_text += f" | Payments: {self._format_currency(total_payment)}"
        
        ctk.CTkLabel(
            header_frame,
            text=stats_text,
            font=ctk.CTkFont(size=11),
            text_color="gray"
        ).pack(side="right", padx=20, pady=10)
        
        # Transaction table
        table_container = tk.Frame(popup, bg='#2b2b2b')
        table_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        txn_columns = ('date', 'description', 'ref_no', 'debit', 'credit', 'balance')
        txn_tree = ttk.Treeview(
            table_container,
            columns=txn_columns,
            show='headings'
        )
        
        txn_col_config = [
            ('date', 'Date', 90),
            ('description', 'Description', 400),
            ('ref_no', 'Reference No.', 120),
            ('debit', 'Debit', 100),
            ('credit', 'Credit', 100),
            ('balance', 'Balance', 120)
        ]
        
        for col_id, heading, width in txn_col_config:
            txn_tree.heading(col_id, text=heading)
            txn_tree.column(col_id, width=width, minwidth=50)
        
        txn_vsb = ttk.Scrollbar(table_container, orient="vertical", command=txn_tree.yview)
        txn_tree.configure(yscrollcommand=txn_vsb.set)
        
        txn_tree.pack(side="left", fill="both", expand=True)
        txn_vsb.pack(side="right", fill="y")
        
        for trans in transactions:
            debit_str = f"{trans['debit']:,.2f}" if trans['debit'] > 0 else ""
            credit_str = f"{trans['credit']:,.2f}" if trans['credit'] > 0 else ""
            
            values = (
                trans['date'][:10] if trans['date'] else "",
                trans['description'][:80] if trans['description'] else "",
                trans['ref_no'][:20] if trans['ref_no'] else "",
                debit_str,
                credit_str,
                trans['balance'][:15] if trans['balance'] else ""
            )
            
            txn_tree.insert('', tk.END, values=values)
        
        ctk.CTkButton(popup, text="Close", command=popup.destroy).pack(pady=10)

    def _export_to_excel(self):
        """Export to Excel"""
        try:
            import tkinter.filedialog as fd
            
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            except ImportError:
                self._show_error("Missing Package", "Please install openpyxl:\npip install openpyxl")
                return
            
            filepath = fd.asksaveasfilename(
                title="Export to Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="Extracted_Accounts.xlsx"
            )
            
            if not filepath:
                return
            
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Account Summary"
            
            headers = [
                "Account Name", "Total Transactions", "Receipt Count", "Payment Count",
                "Total Receipts", "Total Payments", "Net Amount", "Sample Description"
            ]
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            row_num = 2
            for account_name in sorted(self.account_transactions.keys()):
                transactions = self.account_transactions.get(account_name, [])
                
                if not transactions:
                    continue
                
                total_receipt = sum(t['credit'] for t in transactions if t['credit'] > 0)
                total_payment = sum(t['debit'] for t in transactions if t['debit'] > 0)
                receipt_count = sum(1 for t in transactions if t['credit'] > 0)
                payment_count = sum(1 for t in transactions if t['debit'] > 0)
                net_amount = total_receipt - total_payment
                sample_desc = transactions[0]['description'][:100] if transactions[0]['description'] else ""
                
                display_name = self.edited_names.get(account_name, account_name)
                
                values = [
                    display_name,
                    len(transactions),
                    receipt_count,
                    payment_count,
                    total_receipt if total_receipt > 0 else 0,
                    total_payment if total_payment > 0 else 0,
                    net_amount,
                    sample_desc
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws_summary.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    if col in [1, 8]:
                        cell.alignment = Alignment(horizontal="left")
                    elif col in [2, 3, 4]:
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(value, (int, float)) and value != 0:
                            cell.number_format = '#,##0.00'
                
                row_num += 1
            
            # Column widths
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 18
            ws_summary.column_dimensions['C'].width = 15
            ws_summary.column_dimensions['D'].width = 15
            ws_summary.column_dimensions['E'].width = 18
            ws_summary.column_dimensions['F'].width = 18
            ws_summary.column_dimensions['G'].width = 18
            ws_summary.column_dimensions['H'].width = 50
            
            # Transactions sheet
            ws_transactions = wb.create_sheet("All Transactions")
            
            txn_headers = [
                "Account Name", "Date", "Description", "Reference No.",
                "Debit", "Credit", "Balance"
            ]
            
            for col, header in enumerate(txn_headers, 1):
                cell = ws_transactions.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            row_num = 2
            for account_name in sorted(self.account_transactions.keys()):
                transactions = self.account_transactions.get(account_name, [])
                display_name = self.edited_names.get(account_name, account_name)
                
                for trans in transactions:
                    values = [
                        display_name,
                        trans['date'],
                        trans['description'],
                        trans['ref_no'],
                        trans['debit'] if trans['debit'] > 0 else 0,
                        trans['credit'] if trans['credit'] > 0 else 0,
                        trans['balance']
                    ]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws_transactions.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
                        if col in [1, 2, 3, 4, 7]:
                            cell.alignment = Alignment(horizontal="left")
                        else:
                            cell.alignment = Alignment(horizontal="right")
                            if isinstance(value, (int, float)) and value != 0:
                                cell.number_format = '#,##0.00'
                    
                    row_num += 1
            
            ws_transactions.column_dimensions['A'].width = 30
            ws_transactions.column_dimensions['B'].width = 12
            ws_transactions.column_dimensions['C'].width = 50
            ws_transactions.column_dimensions['D'].width = 20
            ws_transactions.column_dimensions['E'].width = 15
            ws_transactions.column_dimensions['F'].width = 15
            ws_transactions.column_dimensions['G'].width = 18
            
            wb.save(filepath)
            
            self._show_success("Export Successful", f"Data exported to:\n{filepath}")
            
        except Exception as e:
            import traceback
            self._show_error("Export Error", str(e) + "\n\n" + traceback.format_exc())

    def _refresh_table(self):
        """Refresh table"""
        self._populate_table(self.search_var.get() if self.search_var else "")

    def _show_success(self, title, message):
        """Show success popup"""
        popup = ctk.CTkToplevel(self)
        popup.title(title)
        popup.geometry("400x150")
        popup.grab_set()
        
        ctk.CTkLabel(
            popup,
            text=message,
            font=ctk.CTkFont(size=12),
            wraplength=350
        ).pack(expand=True, pady=20)
        
        ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)

    def _show_error(self, title, message):
        """Show error popup"""
        popup = ctk.CTkToplevel(self)
        popup.title(title)
        popup.geometry("500x200")
        popup.grab_set()
        
        text_box = ctk.CTkTextbox(popup, width=450, height=100)
        text_box.pack(padx=20, pady=10)
        text_box.insert("1.0", message)
        text_box.configure(state="disabled")
        
        ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)

    def _go_back(self):
        """Go back"""
        print("Going back to name confirmation")
        if self.on_complete_callback:
            self.on_complete_callback(None, None, None, True)

    def _continue(self):
        """Continue to mapping - FIXED"""
        print("Continue to Mapping clicked")
        print(f"Account transactions: {list(self.account_transactions.keys())}")
        print(f"Name aliases: {self.name_aliases}")
        
        if self.on_complete_callback:
            # Get final names from account_transactions (which has resolved names)
            final_names = list(self.account_transactions.keys())
            
            print(f"Calling on_complete_callback with:")
            print(f"  final_names: {final_names}")
            print(f"  name_aliases: {self.name_aliases}")
            
            # Call with: final_names, name_aliases, name_changes, go_back=False
            self.on_complete_callback(final_names, self.name_aliases, {}, False)
        else:
            print("ERROR: on_complete_callback is None!")