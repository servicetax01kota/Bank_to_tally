# ui/new_company_fast_screen.py
"""
New Company Fast Screen - Fast setup for new companies with bulk operations.
"""

import customtkinter as ctk
from typing import List, Dict
import threading


LEDGER_GROUPS = [
    "Sundry Debtors",
    "Sundry Creditors",
    "Bank Accounts",
    "Cash-in-Hand",
    "Direct Expenses",
    "Indirect Expenses",
    "Direct Incomes",
    "Indirect Incomes",
    "Loans & Advances",
    "Capital Account",
    "Current Assets",
    "Current Liabilities",
    "Fixed Assets",
    "Investments",
    "Branch/Divisions",
    "Sales Accounts",
    "Purchase Accounts",
    "Duties & Taxes",
    "Provisions",
    "Suspense A/c"
]


class NewCompanyFastScreen(ctk.CTkFrame):
    """
    Fast setup for new company - Bulk operations.
    All accounts are new (no existing Tally masters to match).
    """
    
    def __init__(self, master, extracted_names: List[str], transactions: List[Dict], on_complete):
        super().__init__(master)
        
        self.extracted_names = list(extracted_names) if extracted_names else []
        self.transactions = list(transactions) if transactions else []
        self.on_complete_callback = on_complete
        
        self.final_mappings = {}
        self.new_ledgers = set()
        self.ledger_groups = {}
        
        self._set_defaults()
        self._create_ui()
        
        print(f"NewCompanyFastScreen initialized with {len(self.extracted_names)} names")

    def _set_defaults(self):
        """Set default groups based on account name patterns"""
        for account_name in self.extracted_names:
            # Default logic based on name patterns
            default_group = "Sundry Debtors"  # Default for receipts (money in)
            
            name_upper = account_name.upper()
            
            # Special cases
            if 'CASH' in name_upper:
                default_group = "Cash-in-Hand"
            elif any(word in name_upper for word in ['CHARGE', 'CHG', 'FEE', 'MANDATE', 'ATM']):
                default_group = "Indirect Expenses"
            elif any(word in name_upper for word in ['BANK', 'HDFC', 'SBI', 'ICICI', 'AXIS', 'PNB']):
                default_group = "Bank Accounts"
            elif any(word in name_upper for word in ['LOAN', 'FINANCE', 'BAJAJ FIN', 'FINTEC']):
                default_group = "Loans & Advances"
            elif any(word in name_upper for word in ['GST', 'TAX', 'TDS', 'CGST', 'SGST', 'IGST']):
                default_group = "Duties & Taxes"
            elif any(word in name_upper for word in ['PAYTM', 'PHONEPE', 'GPAY', 'AMAZON', 'FLIPKART']):
                default_group = "Sundry Creditors"
            
            self.final_mappings[account_name] = account_name
            self.new_ledgers.add(account_name)
            self.ledger_groups[account_name] = default_group

    def _create_ui(self):
        self.grid_rowconfigure(3, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        # Header
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.grid(row=0, column=0, pady=15, padx=20, sticky="ew")
        
        ctk.CTkLabel(
            header_frame,
            text="New Company - Fast Setup",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(side="left")
        
        ctk.CTkLabel(
            header_frame,
            text=f"{len(self.extracted_names)} accounts",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        ).pack(side="right")
        
        # Info
        ctk.CTkLabel(
            self,
            text="Default groups assigned based on account names. Use bulk actions or edit individually.",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        ).grid(row=1, column=0, padx=20, pady=(0, 10), sticky="w")
        
        # Bulk action buttons
        bulk_frame = ctk.CTkFrame(self, fg_color=("gray85", "gray25"))
        bulk_frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        
        ctk.CTkLabel(
            bulk_frame,
            text="Bulk Actions:",
            font=ctk.CTkFont(weight="bold")
        ).pack(side="left", padx=20, pady=10)
        
        ctk.CTkButton(
            bulk_frame,
            text="All to Debtors",
            width=120,
            fg_color="#2196F3",
            command=lambda: self._bulk_set_group("Sundry Debtors")
        ).pack(side="left", padx=5, pady=10)
        
        ctk.CTkButton(
            bulk_frame,
            text="All to Creditors",
            width=120,
            fg_color="#4CAF50",
            command=lambda: self._bulk_set_group("Sundry Creditors")
        ).pack(side="left", padx=5, pady=10)
        
        ctk.CTkButton(
            bulk_frame,
            text="Reset Defaults",
            width=100,
            fg_color="gray",
            command=self._reset_defaults
        ).pack(side="left", padx=5, pady=10)
        
        ctk.CTkButton(
            bulk_frame,
            text="Export for Review",
            width=120,
            fg_color="#9C27B0",
            command=self._export_for_review
        ).pack(side="right", padx=10, pady=10)
        
        # Account list
        self.scroll_frame = ctk.CTkScrollableFrame(self)
        self.scroll_frame.grid(row=3, column=0, padx=20, pady=10, sticky="nsew")
        self.scroll_frame.grid_columnconfigure(0, weight=1)
        
        self._populate_list()
        
        # Loading indicator (hidden initially)
        self.loading_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.loading_label = ctk.CTkLabel(
            self.loading_frame,
            text="Processing... Please wait...",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#FF9800"
        )
        self.loading_label.pack(pady=20)
        
        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.grid(row=4, column=0, padx=20, pady=15, sticky="ew")
        
        ctk.CTkButton(
            btn_frame,
            text="← Back",
            fg_color="gray",
            width=120,
            height=40,
            command=self._go_back
        ).pack(side="left")
        
        ctk.CTkButton(
            btn_frame,
            text="Continue to Review →",
            fg_color="green",
            hover_color="#2E7D32",
            font=ctk.CTkFont(size=13, weight="bold"),
            width=200,
            height=45,
            command=self._continue
        ).pack(side="right")

    def _populate_list(self):
        """Populate account list"""
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()
        
        for i, account_name in enumerate(sorted(self.extracted_names)):
            row = ctk.CTkFrame(
                self.scroll_frame, 
                fg_color=("gray80", "gray20") if i % 2 == 0 else "transparent"
            )
            row.grid(row=i, column=0, padx=5, pady=2, sticky="ew")
            row.grid_columnconfigure(0, weight=1)
            
            # Account name
            name_label = ctk.CTkLabel(row, text=account_name, anchor="w", font=ctk.CTkFont(size=12))
            name_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
            
            # Group dropdown
            group_var = ctk.StringVar(value=self.ledger_groups.get(account_name, "Sundry Debtors"))
            group_dropdown = ctk.CTkComboBox(
                row,
                values=LEDGER_GROUPS,
                variable=group_var,
                width=180,
                command=lambda g, n=account_name: self._update_group(n, g)
            )
            group_dropdown.grid(row=0, column=1, padx=10, pady=5)

    def _update_group(self, account_name, group):
        """Update group for account"""
        self.ledger_groups[account_name] = group

    def _bulk_set_group(self, group):
        """Set all accounts to group"""
        for account_name in self.extracted_names:
            self.ledger_groups[account_name] = group
        self._populate_list()

    def _reset_defaults(self):
        """Reset to defaults"""
        self._set_defaults()
        self._populate_list()
    
    def _export_for_review(self):
        """Export current mapping for review"""
        try:
            import tkinter.filedialog as fd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            filepath = fd.asksaveasfilename(
                title="Export Mapping for Review",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="Account_Mapping_Review.xlsx"
            )
            
            if not filepath:
                return
            
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Account Mapping"
            
            # Headers
            headers = ["Account Name", "Ledger Group", "Corrected Group (Edit this)"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_num, account_name in enumerate(sorted(self.extracted_names), 2):
                ws.cell(row=row_num, column=1, value=account_name)
                ws.cell(row=row_num, column=2, value=self.ledger_groups.get(account_name, ""))
                ws.cell(row=row_num, column=3, value="")  # For user to fill
            
            # Column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            
            # Instructions sheet
            ws_instructions = wb.create_sheet("Instructions")
            
            instructions = [
                "INSTRUCTIONS:",
                "",
                "1. Review the 'Account Mapping' sheet",
                "2. Edit the 'Corrected Group' column (Column C) for any changes",
                "3. Save the file",
                "4. Use 'Import from Review' feature to load changes (if available)",
                "",
                "Available Groups:",
            ]
            
            # Add groups to instructions
            for g in LEDGER_GROUPS:
                instructions.append(f"  - {g}")
            
            for row_num, line in enumerate(instructions, 1):
                cell = ws_instructions.cell(row=row_num, column=1, value=line)
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)
                else:
                    cell.font = Font(size=11)
            
            ws_instructions.column_dimensions['A'].width = 35
            
            wb.save(filepath)
            
            popup = ctk.CTkToplevel(self)
            popup.title("Export Successful")
            popup.geometry("400x120")
            popup.grab_set()
            
            ctk.CTkLabel(
                popup, 
                text=f"Exported to:\n{filepath}",
                text_color="#4CAF50"
            ).pack(expand=True, pady=20)
            
            ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)
            
        except ImportError:
            popup = ctk.CTkToplevel(self)
            popup.title("Error")
            popup.geometry("400x120")
            popup.grab_set()
            
            ctk.CTkLabel(
                popup, 
                text="Please install openpyxl:\npip install openpyxl"
            ).pack(expand=True, pady=20)
            
            ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)
        
        except Exception as e:
            popup = ctk.CTkToplevel(self)
            popup.title("Error")
            popup.geometry("400x150")
            popup.grab_set()
            
            ctk.CTkLabel(
                popup, 
                text=f"Export Error:\n{str(e)}"
            ).pack(expand=True, pady=20)
            
            ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)

    def _go_back(self):
        """Go back"""
        print("NewCompanyFastScreen: Going back")
        if self.on_complete_callback:
            self.on_complete_callback(None, None, None, True)

    def _continue(self):
        """Continue with loading indicator"""
        print("NewCompanyFastScreen: Continue clicked")
        print(f"Final mappings: {self.final_mappings}")
        print(f"New ledgers: {self.new_ledgers}")
        print(f"Ledger groups: {self.ledger_groups}")
        
        # Show loading
        self.loading_frame.grid(row=5, column=0, sticky="ew")
        self.update()
        
        # Process in thread to keep UI responsive
        def process():
            try:
                import time
                time.sleep(0.3)
                self.after(0, self._finish_continue)
            except Exception as e:
                self.after(0, lambda: self._show_error(str(e)))
        
        thread = threading.Thread(target=process)
        thread.start()

    def _finish_continue(self):
        """Finish continue operation"""
        self.loading_frame.grid_remove()
        
        print("NewCompanyFastScreen: Finishing continue")
        print(f"Calling callback with:")
        print(f"  final_mappings: {self.final_mappings}")
        print(f"  new_ledgers: {list(self.new_ledgers)}")
        print(f"  ledger_groups: {self.ledger_groups}")
        
        if self.on_complete_callback:
            self.on_complete_callback(
                self.final_mappings,
                list(self.new_ledgers),
                self.ledger_groups,
                False
            )
        else:
            print("ERROR: on_complete_callback is None!")

    def _show_error(self, message):
        """Show error"""
        self.loading_frame.grid_remove()
        
        popup = ctk.CTkToplevel(self)
        popup.title("Error")
        popup.geometry("400x150")
        popup.grab_set()
        
        ctk.CTkLabel(popup, text=message, wraplength=350).pack(expand=True, pady=20)
        ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)