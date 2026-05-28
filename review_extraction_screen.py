# ui/review_extraction_screen.py
"""
Review Extraction Screen - Allows users to review and correct extracted account names.
"""

import customtkinter as ctk
from typing import List, Dict
import tkinter as tk
from tkinter import ttk
import os


class ReviewExtractionScreen(ctk.CTkFrame):
    """
    Review and correct extracted account names before proceeding.
    """
    
    def __init__(self, master, df, on_complete):
        super().__init__(master)
        
        self.df = df
        self.on_complete_callback = on_complete
        
        # Track corrections: row_index -> corrected_name
        self.corrections = {}
        
        # Get all rows for review
        self.all_rows = self._get_all_rows()
        
        self._create_ui()
        self._populate_table()
    
    def _get_all_rows(self):
        """Get all rows from dataframe for review"""
        rows = []
        
        for idx, row in self.df.iterrows():
            extracted = row.get('EXTRACTED_NAME', '')
            final_name = row.get('FINAL_NAME', '')
            description = str(row.get('DESCRIPTION', ''))
            debit = float(row.get('DEBIT', 0) or 0)
            credit = float(row.get('CREDIT', 0) or 0)
            trans_type = row.get('TRANS_TYPE', 'Payment')
            
            needs_review = (
                not extracted or 
                extracted in ['Payment made', 'Payment Receipt', None, '', 'None'] or
                final_name in ['Payment made', 'Payment Receipt']
            )
            
            rows.append({
                'row_index': idx,
                'description': description[:150],
                'extracted_name': extracted if extracted else '-',
                'final_name': final_name,
                'debit': debit,
                'credit': credit,
                'trans_type': trans_type,
                'needs_review': needs_review
            })
        
        return rows
    
    def _create_ui(self):
        self.grid_rowconfigure(3, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        # Header
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.grid(row=0, column=0, pady=15, padx=20, sticky="ew")
        
        ctk.CTkLabel(
            header_frame,
            text="Review Extracted Account Names",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(side="left")
        
        review_count = sum(1 for r in self.all_rows if r['needs_review'])
        total_count = len(self.all_rows)
        
        stats_text = f"{total_count} transactions"
        if review_count > 0:
            stats_text += f" | {review_count} need review"
        
        self.stats_label = ctk.CTkLabel(
            header_frame,
            text=stats_text,
            font=ctk.CTkFont(size=12),
            text_color="#FF9800" if review_count > 0 else "#4CAF50"
        )
        self.stats_label.pack(side="right")
        
        # Filter toolbar
        filter_frame = ctk.CTkFrame(self, fg_color=("gray85", "gray25"))
        filter_frame.grid(row=1, column=0, padx=20, pady=(0, 5), sticky="ew")
        
        ctk.CTkLabel(filter_frame, text="Filter:").pack(side="left", padx=10)
        
        self.filter_var = ctk.StringVar(value="All")
        filter_dropdown = ctk.CTkComboBox(
            filter_frame,
            values=["All", "Needs Review", "Has Name", "Cash", "Bank Charges", "Payment made", "Payment Receipt"],
            variable=self.filter_var,
            width=150,
            command=self._apply_filter
        )
        filter_dropdown.pack(side="left", padx=5)
        
        ctk.CTkLabel(filter_frame, text="Search:").pack(side="left", padx=(20, 5))
        
        self.search_var = ctk.StringVar()
        search_entry = ctk.CTkEntry(filter_frame, textvariable=self.search_var, width=200)
        search_entry.pack(side="left", padx=5)
        search_entry.bind('<KeyRelease>', self._apply_filter)
        
        # Bulk actions
        bulk_frame = ctk.CTkFrame(self, fg_color=("gray85", "gray25"))
        bulk_frame.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="ew")
        
        ctk.CTkLabel(bulk_frame, text="Bulk Actions:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
        
        ctk.CTkButton(
            bulk_frame, text="Set Selected To:", width=110, fg_color="#2196F3", command=self._bulk_set_selected
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            bulk_frame, text="Set All 'Payment made' to:", width=160, fg_color="#FF9800", command=self._bulk_set_payment_made
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            bulk_frame, text="Export to Excel", width=110, fg_color="#4CAF50", command=self._export_to_excel
        ).pack(side="right", padx=10)
        
        ctk.CTkButton(
            bulk_frame, text="Import from Excel", width=110, fg_color="#9C27B0", command=self._import_from_excel
        ).pack(side="right", padx=5)
        
        # Table
        self._create_table()
        
        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.grid(row=4, column=0, padx=20, pady=15, sticky="ew")
        
        ctk.CTkButton(btn_frame, text="Back", fg_color="gray", command=self._go_back).pack(side="left")
        
        correction_count = len(self.corrections)
        continue_text = "Confirm & Continue →"
        if correction_count > 0:
            continue_text += f" ({correction_count} corrections)"
        
        ctk.CTkButton(
            btn_frame, text=continue_text, fg_color="green", font=ctk.CTkFont(size=13, weight="bold"), command=self._continue
        ).pack(side="right")
    
    def _create_table(self):
        """Create treeview table"""
        table_frame = ctk.CTkFrame(self)
        table_frame.grid(row=3, column=0, padx=20, pady=5, sticky="nsew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        tree_container = tk.Frame(table_frame, bg='#2b2b2b')
        tree_container.grid(row=0, column=0, sticky="nsew")
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        columns = ('select', 'date', 'trans_type', 'description', 'extracted_name', 'final_name', 'debit', 'credit', 'status')
        
        self.tree = ttk.Treeview(tree_container, columns=columns, show='headings', selectmode='extended')
        
        col_config = [
            ('select', '☐', 40, tk.CENTER),
            ('date', 'Date', 100, tk.CENTER),
            ('trans_type', 'Type', 70, tk.CENTER),
            ('description', 'Description', 280, tk.W),
            ('extracted_name', 'Extracted', 120, tk.W),
            ('final_name', 'Final Name (Double-click)', 150, tk.W),
            ('debit', 'Debit', 90, tk.E),
            ('credit', 'Credit', 90, tk.E),
            ('status', 'Status', 80, tk.CENTER)
        ]
        
        for col_id, heading, width, anchor in col_config:
            self.tree.heading(col_id, text=heading, anchor=anchor)
            self.tree.column(col_id, width=width, anchor=anchor, minwidth=30)
        
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        style = ttk.Style()
        if 'clam' in style.theme_names():
            style.theme_use('clam')
        
        style.configure("Treeview", background="#2d2d2d", foreground="#e0e0e0", fieldbackground="#2d2d2d", font=('Segoe UI', 9), rowheight=25)
        style.configure("Treeview.Heading", background="#4a4a4a", foreground="white", font=('Segoe UI', 9, 'bold'))
        style.map('Treeview', background=[('selected', '#1976D2')], foreground=[('selected', 'white')])
        
        self.tree.bind('<Double-1>', self._on_double_click)
        
        self.status_label = ctk.CTkLabel(table_frame, text="", font=ctk.CTkFont(size=10), text_color="gray")
        self.status_label.grid(row=2, column=0, sticky="w", pady=5)
    
    def _populate_table(self, filter_text=None, filter_type="All"):
        """Populate table with data"""
        if self.tree is None:
            return
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        row_num = 0
        for row_data in self.all_rows:
            if filter_type == "Needs Review" and not row_data['needs_review']:
                continue
            elif filter_type == "Has Name" and row_data['needs_review']:
                continue
            elif filter_type in ["Cash", "Bank Charges", "Payment made", "Payment Receipt"]:
                if row_data['final_name'] != filter_type:
                    continue
            
            if filter_text:
                search_lower = filter_text.lower()
                if (search_lower not in row_data['description'].lower() and 
                    search_lower not in str(row_data['extracted_name']).lower() and
                    search_lower not in row_data['final_name'].lower()):
                    continue
            
            if row_data['needs_review']:
                status = "⚠️ Review"
                status_tag = 'review'
            elif row_data['final_name'] in ['Cash', 'Bank Charges']:
                status = "✓ Auto"
                status_tag = 'auto'
            else:
                status = "✓ OK"
                status_tag = 'ok'
            
            # Get date from dataframe
            date_val = ""
            try:
                actual_idx = row_data['row_index']
                if 'DATE' in self.df.columns:
                    date_raw = self.df.at[actual_idx, 'DATE']
                    date_val = str(date_raw)[:10] if date_raw else ""
            except:
                pass
            
            debit_str = f"{row_data['debit']:,.2f}" if row_data['debit'] > 0 else ""
            credit_str = f"{row_data['credit']:,.2f}" if row_data['credit'] > 0 else ""
            
            values = ("☐", date_val, row_data['trans_type'], row_data['description'][:100],
                     row_data['extracted_name'], row_data['final_name'], debit_str, credit_str, status)
            
            tag = 'even' if row_num % 2 == 0 else 'odd'
            self.tree.insert('', tk.END, values=values, tags=(tag, status_tag))
            row_num += 1
        
        self.tree.tag_configure('even', background='#2d2d2d')
        self.tree.tag_configure('odd', background='#353535')
        self.tree.tag_configure('review', foreground='#FF9800')
        self.tree.tag_configure('auto', foreground='#4CAF50')
        self.tree.tag_configure('ok', foreground='#e0e0e0')
        self.tree.tag_configure('edited', foreground='#2196F3')
        
        self._update_status()
    
    def _apply_filter(self, event=None):
        filter_type = self.filter_var.get()
        search_text = self.search_var.get()
        self._populate_table(search_text, filter_type)
    
    def _on_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        selected = self.tree.selection()
        if not selected:
            return
        
        item = selected[0]
        column = self.tree.identify_column(event.x)
        
        if column != '#6':
            return
        
        values = self.tree.item(item, 'values')
        current_name = values[5]
        
        self._edit_name_popup(item, current_name)
    
    def _edit_name_popup(self, item, current_name):
        popup = ctk.CTkToplevel(self)
        popup.title("Edit Account Name")
        popup.geometry("450x220")
        popup.grab_set()
        
        ctk.CTkLabel(popup, text="Enter correct account name:", font=ctk.CTkFont(size=12)).pack(pady=(20, 10))
        
        name_var = ctk.StringVar(value=current_name)
        entry = ctk.CTkEntry(popup, textvariable=name_var, width=400, font=ctk.CTkFont(size=12))
        entry.pack(padx=20, pady=10)
        entry.focus()
        entry.select_range(0, 'end')
        
        quick_frame = ctk.CTkFrame(popup, fg_color="transparent")
        quick_frame.pack(pady=5)
        
        for name in ['Cash', 'Bank Charges', 'Payment made', 'Payment Receipt']:
            ctk.CTkButton(quick_frame, text=name, width=90, fg_color="gray", command=lambda n=name: name_var.set(n)).pack(side="left", padx=2)
        
        def save():
            new_name = name_var.get().strip()
            if new_name:
                values = list(self.tree.item(item, 'values'))
                values[5] = new_name
                values[8] = "✓ Edited"
                self.tree.item(item, values=values, tags=('edited',))
                
                row_idx = self.tree.index(item)
                if row_idx < len(self.all_rows):
                    actual_row = self.all_rows[row_idx]['row_index']
                    self.corrections[actual_row] = new_name
                
                popup.destroy()
                self._update_status()
        
        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(pady=10)
        ctk.CTkButton(btn_frame, text="Save", fg_color="green", width=80, command=save).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Cancel", fg_color="gray", width=80, command=popup.destroy).pack(side="left", padx=10)
        entry.bind('<Return>', lambda e: save())
    
    def _bulk_set_selected(self):
        selected = self.tree.selection()
        if not selected:
            self._show_message("No Selection", "Please select rows to bulk edit.")
            return
        self._bulk_set_popup(f"Set account name for {len(selected)} selected rows:", selected)
    
    def _bulk_set_payment_made(self):
        payment_made_items = []
        for item in self.tree.get_children():
            values = self.tree.item(item, 'values')
            if values[5] == 'Payment made':
                payment_made_items.append(item)
        
        if not payment_made_items:
            self._show_message("No Entries", "No 'Payment made' entries found.")
            return
        
        self.tree.selection_set(payment_made_items)
        self._bulk_set_popup(f"Set account name for {len(payment_made_items)} 'Payment made' entries:", payment_made_items)
    
    def _bulk_set_popup(self, message, items):
        popup = ctk.CTkToplevel(self)
        popup.title("Bulk Set Account Name")
        popup.geometry("450x200")
        popup.grab_set()
        
        ctk.CTkLabel(popup, text=message, font=ctk.CTkFont(size=12)).pack(pady=(20, 10))
        
        name_var = ctk.StringVar()
        entry = ctk.CTkEntry(popup, textvariable=name_var, width=400, font=ctk.CTkFont(size=12))
        entry.pack(padx=20, pady=10)
        entry.focus()
        
        quick_frame = ctk.CTkFrame(popup, fg_color="transparent")
        quick_frame.pack(pady=5)
        
        for name in ['Cash', 'Bank Charges', 'Payment made', 'Payment Receipt']:
            ctk.CTkButton(quick_frame, text=name, width=90, fg_color="gray", command=lambda n=name: name_var.set(n)).pack(side="left", padx=2)
        
        def apply():
            new_name = name_var.get().strip()
            if new_name:
                for item in items:
                    values = list(self.tree.item(item, 'values'))
                    values[5] = new_name
                    values[8] = "✓ Edited"
                    self.tree.item(item, values=values, tags=('edited',))
                    
                    row_idx = self.tree.index(item)
                    if row_idx < len(self.all_rows):
                        actual_row = self.all_rows[row_idx]['row_index']
                        self.corrections[actual_row] = new_name
                
                popup.destroy()
                self._update_status()
        
        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(pady=10)
        ctk.CTkButton(btn_frame, text="Apply", fg_color="green", width=80, command=apply).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Cancel", fg_color="gray", width=80, command=popup.destroy).pack(side="left", padx=10)
        entry.bind('<Return>', lambda e: apply())
    
    def _export_to_excel(self):
        """Export review data to Excel"""
        try:
            import tkinter.filedialog as fd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            filepath = fd.asksaveasfilename(
                title="Export to Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="Review_Extraction.xlsx"
            )
            
            if not filepath:
                return
            
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Review Extraction"
            
            headers = ["Row #", "Date", "Type", "Description", "Debit", "Credit", "Extracted Name", "Final Name", "Corrected Name", "Status"]
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for row_num, row_data in enumerate(self.all_rows, 2):
                actual_idx = row_data['row_index']
                
                date_val = ""
                try:
                    if 'DATE' in self.df.columns:
                        date_raw = self.df.at[actual_idx, 'DATE']
                        date_val = str(date_raw)[:10] if date_raw else ""
                except:
                    pass
                
                corrected_name = self.corrections.get(actual_idx, "")
                
                if corrected_name:
                    status = "Corrected"
                elif row_data['needs_review']:
                    status = "Needs Review"
                else:
                    status = "OK"
                
                values = [
                    row_num - 1, date_val, row_data['trans_type'], row_data['description'],
                    row_data['debit'] if row_data['debit'] > 0 else 0,
                    row_data['credit'] if row_data['credit'] > 0 else 0,
                    row_data['extracted_name'], row_data['final_name'], corrected_name, status
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    
                    if col == 10:
                        if status == "Corrected":
                            cell.font = Font(color="2196F3")
                        elif status == "Needs Review":
                            cell.font = Font(color="FF9800")
                        else:
                            cell.font = Font(color="4CAF50")
                    
                    if col in [1, 2, 3, 10]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col in [5, 6]:
                        cell.alignment = Alignment(horizontal="right")
                        if value != 0:
                            cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = Alignment(horizontal="left")
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 12
            
            wb.save(filepath)
            
            self._show_success("Export Successful", f"Data exported to:\n{filepath}\n\nTotal rows: {len(self.all_rows)}\nNeed review: {sum(1 for r in self.all_rows if r['needs_review'])}")
            
        except ImportError:
            self._show_error("Missing Package", "Please install openpyxl:\npip install openpyxl")
        except Exception as e:
            import traceback
            self._show_error("Export Error", str(e) + "\n\n" + traceback.format_exc())
    
    def _import_from_excel(self):
        """Import corrections from Excel file"""
        try:
            import tkinter.filedialog as fd
            from openpyxl import load_workbook
            
            filepath = fd.askopenfilename(title="Import Corrections from Excel", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            
            if not filepath:
                return
            
            wb = load_workbook(filepath, data_only=True)
            
            if "Review Extraction" in wb.sheetnames:
                ws = wb["Review Extraction"]
            else:
                ws = wb.active
            
            corrections_imported = 0
            headers = [cell.value for cell in ws[1]]
            
            try:
                corrected_name_col = headers.index("Corrected Name") + 1
            except ValueError:
                corrected_name_col = 9
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if len(row) < corrected_name_col:
                    continue
                
                corrected_name = row[corrected_name_col - 1]
                
                if corrected_name and str(corrected_name).strip():
                    row_idx_value = row[0]
                    
                    if isinstance(row_idx_value, int) and 1 <= row_idx_value <= len(self.all_rows):
                        actual_idx = self.all_rows[row_idx_value - 1]['row_index']
                        self.corrections[actual_idx] = str(corrected_name).strip()
                        corrections_imported += 1
            
            self._populate_table(self.search_var.get(), self.filter_var.get())
            
            self._show_success("Import Successful", f"Imported {corrections_imported} corrections from:\n{filepath}")
            
        except ImportError:
            self._show_error("Missing Package", "Please install openpyxl:\npip install openpyxl")
        except Exception as e:
            import traceback
            self._show_error("Import Error", str(e) + "\n\n" + traceback.format_exc())
    
    def _update_status(self):
        total = len(self.tree.get_children())
        review = sum(1 for r in self.all_rows if r['needs_review'])
        edited = len(self.corrections)
        self.status_label.configure(text=f"Showing {total} | {review} need review | {edited} corrections")
    
    def _go_back(self):
        if self.on_complete_callback:
            self.on_complete_callback(None, True)
    
    def _continue(self):
        for row_idx, new_name in self.corrections.items():
            self.df.at[row_idx, 'FINAL_NAME'] = new_name
            self.df.at[row_idx, 'EXTRACTED_NAME'] = new_name
        
        if self.on_complete_callback:
            self.on_complete_callback(self.df, False)
    
    def _show_message(self, title, message):
        popup = ctk.CTkToplevel(self)
        popup.title(title)
        popup.geometry("350x120")
        popup.grab_set()
        ctk.CTkLabel(popup, text=message, wraplength=300).pack(expand=True, pady=20)
        ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)
    
    def _show_success(self, title, message):
        popup = ctk.CTkToplevel(self)
        popup.title(title)
        popup.geometry("450x180")
        popup.grab_set()
        ctk.CTkLabel(popup, text=message, font=ctk.CTkFont(size=11), text_color="#4CAF50", wraplength=400, justify="left").pack(expand=True, pady=20)
        ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)
    
    def _show_error(self, title, message):
        popup = ctk.CTkToplevel(self)
        popup.title(title)
        popup.geometry("500x250")
        popup.grab_set()
        text_box = ctk.CTkTextbox(popup, width=450, height=150)
        text_box.pack(padx=20, pady=10)
        text_box.insert("1.0", message)
        text_box.configure(state="disabled")
        ctk.CTkButton(popup, text="OK", command=popup.destroy).pack(pady=10)